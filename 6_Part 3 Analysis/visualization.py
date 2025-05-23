import pandas as pd
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import geopandas as gpd
from adjustText import adjust_text

# === Hauptpfade ===
BASE_FOLDER = "7_Part 4 Graphics and tables"
CSV_FOLDER = os.path.join(BASE_FOLDER, "csv-files")
XLSX_FOLDER = os.path.join(BASE_FOLDER, "xlsx-files")
PNG_FOLDER = os.path.join(BASE_FOLDER, "png-files")

# === Pfade für Heatmaps ===
SHAPEFILE_PATH = os.path.join("2_Shapefile folder", "Stadtteile_Karlsruhe.shp")
HEATMAP_STADTBEZIEHUNGEN_PATH = os.path.join(PNG_FOLDER, "Stadtteilbeziehungen_Heatmap.png")
HEATMAP_ZIELE_PATH = os.path.join(PNG_FOLDER, "Beliebte Arbeits- und Freizeitziele.png")

# === Eingabepfad ===
CSV_PATH = os.path.join("3_Data for analysis", "wegetagebuch_karlsruhe_koordinaten.csv")

# === Alle Ausgabepfade ===
OUTPUT_CSV_PATH = os.path.join(CSV_FOLDER, "Stadtteilbeziehungen_Wegeanzahl.csv")
OUTPUT_XLSX_PATH = os.path.join(XLSX_FOLDER, "Stadtteilbeziehungen_Wegeanzahl.xlsx")
RANKING_CSV_PATH = os.path.join(CSV_FOLDER, "Stadtteile_Ranking.csv")
RANKING_XLSX_PATH = os.path.join(XLSX_FOLDER, "Stadtteile_Ranking.xlsx")
ZWECK_CSV_PATH = os.path.join(CSV_FOLDER, "Verkehrsaufkommen (Wege)_Wegegründe.csv")
ZWECK_XLSX_PATH = os.path.join(XLSX_FOLDER, "Verkehrsaufkommen (Wege)_Wegegründe.xlsx")
ZWECK_PIE_PATH = os.path.join(PNG_FOLDER, "Verkehrsaufkommen (Wege)_Wegegründe_PieChart.png")
MULTIMODAL_CSV_PATH = os.path.join(CSV_FOLDER, "Multimodalität.csv")
MULTIMODAL_XLSX_PATH = os.path.join(XLSX_FOLDER, "Multimodalität.xlsx")
MODAL_SPLIT_CSV_PATH = os.path.join(CSV_FOLDER, "Modal Split_Wege.csv")
MODAL_SPLIT_XLSX_PATH = os.path.join(XLSX_FOLDER, "Modal Split_Wege.xlsx")
MODAL_SPLIT_PIE_PATH = os.path.join(PNG_FOLDER, "Modal Split_Wege.png")
MODAL_SPLIT_KM_CSV_PATH = os.path.join(CSV_FOLDER, "Modal Split_Personenkilometer.csv")
MODAL_SPLIT_KM_XLSX_PATH = os.path.join(XLSX_FOLDER, "Modal Split_Personenkilometer.xlsx")
MODAL_SPLIT_KM_PIE_PATH = os.path.join(PNG_FOLDER, "Modal Split_Personenkilometer.png")
WEGE_DISTANZEN_CSV_PATH = os.path.join(CSV_FOLDER, "Wege und Distanzen.csv")
WEGE_DISTANZEN_XLSX_PATH = os.path.join(XLSX_FOLDER, "Wege und Distanzen.xlsx")
WOHNVIERTEL_CSV_PATH = os.path.join(CSV_FOLDER, "Wohnviertel.csv")
WOHNVIERTEL_XLSX_PATH = os.path.join(XLSX_FOLDER, "Wohnviertel.xlsx")
HAUSHALTSTYPEN_CSV_PATH = os.path.join(CSV_FOLDER, "Haushaltstypen.csv")
HAUSHALTSTYPEN_XLSX_PATH = os.path.join(XLSX_FOLDER, "Haushaltstypen.xlsx")
BERUF_CSV_PATH = os.path.join(CSV_FOLDER, "Berufe.csv")
BERUF_XLSX_PATH = os.path.join(XLSX_FOLDER, "Berufe.xlsx")
GESCHLECHTER_CSV_PATH = os.path.join(CSV_FOLDER, "Geschlechter.csv")
GESCHLECHTER_XLSX_PATH = os.path.join(XLSX_FOLDER, "Geschlechter.xlsx")
ALTER_CSV_PATH = os.path.join(CSV_FOLDER, "Durchschnittsalter.csv")
ALTER_XLSX_PATH = os.path.join(XLSX_FOLDER, "Durchschnittsalter.xlsx")
WEGE_ANZAHL_CSV_PATH = os.path.join(CSV_FOLDER, "Anzahl aufgezeichneter Wege.csv")
WEGE_ANZAHL_XLSX_PATH = os.path.join(XLSX_FOLDER, "Anzahl aufgezeichneter Wege.xlsx")
PERSONEN_ANZAHL_CSV_PATH = os.path.join(CSV_FOLDER, "Anzahl befragter Personen.csv")
PERSONEN_ANZAHL_XLSX_PATH = os.path.join(XLSX_FOLDER, "Anzahl befragter Personen.xlsx")
SPITZENSTUNDE_CSV_PATH = os.path.join(CSV_FOLDER, "Spitzenstunde.csv")
SPITZENSTUNDE_XLSX_PATH = os.path.join(XLSX_FOLDER, "Spitzenstunde.xlsx")

# === Hilfsfunktion: Excel-Dateien formatieren ===
def format_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2
        for cell in column_cells:
            cell.alignment = Alignment(horizontal='center')
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)

# === Daten einlesen ===
df = pd.read_csv(CSV_PATH)

# === Vorbereitungen: Startzeit und Datum ===
df["Startzeit"] = pd.to_datetime(df["Startzeit"], errors='coerce')
df["Datum"] = df["Startzeit"].dt.date

# === 1. Stadtteilbeziehungen ===
df_valid = df.dropna(subset=["Stadtteil Start", "Stadtteil Ziel"])
weg_counts = (
    df_valid.groupby(["Stadtteil Start", "Stadtteil Ziel"])
    .size()
    .reset_index(name="Anzahl Wege")
    .sort_values(by="Anzahl Wege", ascending=False)
)
mask_häufig = weg_counts["Anzahl Wege"] >= 20
weg_counts_häufig = weg_counts[mask_häufig]
weg_counts_selten = weg_counts[~mask_häufig]
restliche_wege_anzahl = weg_counts_selten["Anzahl Wege"].sum()

output_df = weg_counts_häufig.copy()
if restliche_wege_anzahl > 0:
    output_df = pd.concat(
        [
            output_df,
            pd.DataFrame({
                "Stadtteil Start": ["Restliche Wege"],
                "Stadtteil Ziel": [""],
                "Anzahl Wege": [restliche_wege_anzahl],
            }),
        ],
        ignore_index=True,
    )

output_df.to_csv(OUTPUT_CSV_PATH, index=False, encoding="utf-8-sig")
output_df.to_excel(OUTPUT_XLSX_PATH, index=False)
format_excel(OUTPUT_XLSX_PATH)

# === 1.1 Heatmap der Stadtteilbeziehungen ===
try:
    gdf = gpd.read_file(SHAPEFILE_PATH)
except Exception as e:
    print(f"❗ Fehler beim Laden der Shapedatei: {e}")
    gdf = None

if gdf is not None:
    # Stadtteilnamen und Zentroiden bestimmen
    string_cols = gdf.select_dtypes(include=["object"]).columns.tolist()
    name_col_candidates = [c for c in string_cols if "name" in c.lower() or "stadtteil" in c.lower()]
    name_col = name_col_candidates[0] if name_col_candidates else string_cols[0]
    gdf["district_name"] = gdf[name_col]
    gdf["centroid"] = gdf.geometry.centroid
    centroid_dict = dict(zip(gdf["district_name"], gdf["centroid"]))

    # Stadtteilbeziehungen einlesen
    rel_df = output_df.query("`Stadtteil Start` != 'Restliche Wege' and `Stadtteil Ziel` != ''")
    if not rel_df.empty:
        flow_start = rel_df.groupby("Stadtteil Start")["Anzahl Wege"].sum()
        flow_target = rel_df.groupby("Stadtteil Ziel")["Anzahl Wege"].sum()
        gdf["flow_sum"] = gdf["district_name"].map(flow_start.add(flow_target, fill_value=0)).fillna(0)
        max_count = rel_df["Anzahl Wege"].max()

        fig, ax = plt.subplots(figsize=(12, 12))
        gdf.plot(ax=ax, column="flow_sum", cmap="Reds", edgecolor="black", linewidth=0.5)

        for _, row in rel_df.iterrows():
            start = row["Stadtteil Start"]
            ziel = row["Stadtteil Ziel"]
            count = row["Anzahl Wege"]
            if start in centroid_dict and ziel in centroid_dict:
                x0, y0 = centroid_dict[start].x, centroid_dict[start].y
                x1, y1 = centroid_dict[ziel].x, centroid_dict[ziel].y
                width = 1 + (count / max_count) * 8
                ax.annotate(
                    "",
                    xy=(x1, y1),
                    xytext=(x0, y0),
                    arrowprops=dict(arrowstyle="->", color="darkred", linewidth=width, alpha=0.7),
                )

        for _, row in gdf.iterrows():
            ax.text(row["centroid"].x, row["centroid"].y, row["district_name"], ha="center", fontsize=8)

        ax.set_title("Heatmap der Stadtteilbeziehungen in Karlsruhe")
        ax.axis("off")
        plt.savefig(HEATMAP_STADTBEZIEHUNGEN_PATH, bbox_inches="tight")
        plt.close()

# === 1.2 Karte beliebter Arbeits- und Freizeitziele ===
def get_popular(df_sub):
    counts = df_sub["Zielort"].value_counts()
    return counts[counts >= 3]

arbeit_pop = get_popular(df.query("Zweck == 'Arbeit' and Zielort.notna() and `Stadtteil Ziel`.notna()"))
frei_pop = get_popular(df.query("Zweck == 'Freizeit' and Zielort.notna() and `Stadtteil Ziel`.notna()"))

if arbeit_pop.empty and frei_pop.empty:
    print("ℹ️ Keine häufigen Arbeits- oder Freizeitziele gefunden.")
else:
    fig2, ax2 = plt.subplots(figsize=(12, 12))
    gdf.plot(ax=ax2, facecolor="none", edgecolor="black", linewidth=0.5)

    # Arbeitsziele
    for ziel, cnt in arbeit_pop.items():
        district = df.loc[df["Zielort"] == ziel, "Stadtteil Ziel"].iloc[0]
        if district in centroid_dict:
            pt = centroid_dict[district]
            ax2.scatter(pt.x, pt.y, color="red", s=80, label="Arbeitsziel" if "Arbeitsziel" not in ax2.get_legend_handles_labels()[1] else "", zorder=5)
            ax2.text(pt.x, pt.y, f"{ziel} ({int(cnt)}-Mal)", fontsize=8, color="darkred", ha='center', va='bottom')

    # Freizeitziele
    for ziel, cnt in frei_pop.items():
        district = df.loc[df["Zielort"] == ziel, "Stadtteil Ziel"].iloc[0]
        if district in centroid_dict:
            pt = centroid_dict[district]
            ax2.scatter(pt.x, pt.y, color="green", s=80, label="Freizeitziel" if "Freizeitziel" not in ax2.get_legend_handles_labels()[1] else "", zorder=5)
            ax2.text(pt.x, pt.y, f"{ziel} ({int(cnt)}-Mal)", fontsize=8, color="darkgreen", ha='center', va='bottom')

    ax2.legend(loc="upper right")
    ax2.set_title("Karte beliebter Arbeits- und Freizeitziele")
    ax2.axis("off")
    plt.savefig(HEATMAP_ZIELE_PATH, bbox_inches="tight")
    plt.close()

# === 2. Stadtteile-Ranking ===
start_ranking = df_valid["Stadtteil Start"].value_counts().reset_index()
start_ranking.columns = ["Start Stadtteil", "Anzahl Starts"]
ziel_ranking = df_valid["Stadtteil Ziel"].value_counts().reset_index()
ziel_ranking.columns = ["Ziel Stadtteil", "Anzahl Ziele"]

max_len = max(len(start_ranking), len(ziel_ranking))
start_ranking = start_ranking.reindex(range(max_len))
ziel_ranking = ziel_ranking.reindex(range(max_len))

ranking_df = pd.DataFrame({
    "Start Stadtteil": start_ranking["Start Stadtteil"],
    "Anzahl Starts": start_ranking["Anzahl Starts"],
    "Ziel Stadtteil": ziel_ranking["Ziel Stadtteil"],
    "Anzahl Ziele": ziel_ranking["Anzahl Ziele"]
})

ranking_df.to_csv(RANKING_CSV_PATH, index=False, encoding='utf-8-sig')
ranking_df.to_excel(RANKING_XLSX_PATH, index=False)
format_excel(RANKING_XLSX_PATH)

# === 3. Wegezwecke Analyse
zweck_valid = df.dropna(subset=["Zweck"])
zweck_counts = (
    zweck_valid["Zweck"].value_counts(normalize=True) * 100
).reset_index()
zweck_counts.columns = ["Zweck", "Prozentuale Verteilung"]
zweck_counts["Prozentuale Verteilung"] = zweck_counts["Prozentuale Verteilung"].round(2)

zweck_counts.to_csv(ZWECK_CSV_PATH, index=False, encoding='utf-8-sig')
zweck_counts.to_excel(ZWECK_XLSX_PATH, index=False)
format_excel(ZWECK_XLSX_PATH)

farben_dict_zwecke = {
    "Heimweg": "orchid", "Sport": "limegreen", "Schule/Uni": "navy",
    "Freizeit": "gold", "Arbeit": "deepskyblue", "Arztbesuch": "grey",
    "Begleitung": "lightgreen", "Erholung": "sandybrown", "Einkaufen": "darkred"
}
farben = [farben_dict_zwecke.get(zweck, "lightgrey") for zweck in zweck_counts["Zweck"]]

fig, ax = plt.subplots(figsize=(10, 8))
ax.pie(zweck_counts["Prozentuale Verteilung"], labels=zweck_counts["Zweck"], autopct="%1.1f%%", startangle=140, colors=farben)
ax.axis('equal')
plt.title("Prozentuale Verteilung der Wegezwecke")
plt.savefig(ZWECK_PIE_PATH, bbox_inches='tight')
plt.close()

# === 4. Multimodalität Analyse
multimodal_valid = df.dropna(subset=["Multimodal"])
multimodal_ja = (multimodal_valid["Multimodal"].str.lower() == "ja").sum()
gesamt = multimodal_valid.shape[0]
multimodal_prozent = round((multimodal_ja / gesamt) * 100, 2)
monomodal_prozent = round(100 - multimodal_prozent, 2)

multimodal_df = pd.DataFrame({
    "Typ": ["Multimodal", "Monomodal"],
    "Prozentuale Verteilung": [multimodal_prozent, monomodal_prozent]
})

multimodal_df.to_csv(MULTIMODAL_CSV_PATH, index=False, encoding='utf-8-sig')
multimodal_df.to_excel(MULTIMODAL_XLSX_PATH, index=False)
format_excel(MULTIMODAL_XLSX_PATH)

# === 5. Modal Split Wege
modal_split_valid = df.dropna(subset=["Verkehrsmittel"])
modal_split = (
    modal_split_valid["Verkehrsmittel"].value_counts(normalize=True) * 100
).reset_index()
modal_split.columns = ["Verkehrsmittel", "Prozentuale Verteilung"]
modal_split["Prozentuale Verteilung"] = modal_split["Prozentuale Verteilung"].round(2)

modal_split.to_csv(MODAL_SPLIT_CSV_PATH, index=False, encoding='utf-8-sig')
modal_split.to_excel(MODAL_SPLIT_XLSX_PATH, index=False)
format_excel(MODAL_SPLIT_XLSX_PATH)

farben_dict_modal = {
    "Auto": "red", "zu Fuß": "deepskyblue", "Fahrrad": "navy",
    "ÖPNV": "green", "Multimodal": "orange", "E-Scooter": "purple"
}
farben_modal = [farben_dict_modal.get(vm, "lightgrey") for vm in modal_split["Verkehrsmittel"]]

fig, ax = plt.subplots(figsize=(10, 8))
ax.pie(modal_split["Prozentuale Verteilung"], labels=modal_split["Verkehrsmittel"], autopct="%1.1f%%", startangle=140, colors=farben_modal)
ax.axis('equal')
plt.title("Modal Split der Wege")
plt.savefig(MODAL_SPLIT_PIE_PATH, bbox_inches='tight')
plt.close()

# === 6. Modal Split Personenkilometer
modal_split_km = df.dropna(subset=["Verkehrsmittel", "Entfernung_km"])
modal_split_km_grouped = modal_split_km.groupby("Verkehrsmittel")["Entfernung_km"].sum()
modal_split_km_percent = (modal_split_km_grouped / modal_split_km_grouped.sum() * 100).reset_index()
modal_split_km_percent.columns = ["Verkehrsmittel", "Prozentuale Verteilung"]
modal_split_km_percent["Prozentuale Verteilung"] = modal_split_km_percent["Prozentuale Verteilung"].round(2)

modal_split_km_percent.to_csv(MODAL_SPLIT_KM_CSV_PATH, index=False, encoding='utf-8-sig')
modal_split_km_percent.to_excel(MODAL_SPLIT_KM_XLSX_PATH, index=False)
format_excel(MODAL_SPLIT_KM_XLSX_PATH)

farben_modal_km = [farben_dict_modal.get(vm, "lightgrey") for vm in modal_split_km_percent["Verkehrsmittel"]]

fig, ax = plt.subplots(figsize=(10, 8))
ax.pie(modal_split_km_percent["Prozentuale Verteilung"], labels=modal_split_km_percent["Verkehrsmittel"], autopct="%1.1f%%", startangle=140, colors=farben_modal_km)
ax.axis('equal')
plt.title("Modal Split der Personenkilometer")
plt.savefig(MODAL_SPLIT_KM_PIE_PATH, bbox_inches='tight')
plt.close()

# === 7. Wege und Distanzen
df_filtered = df[["PersonenID", "Datum", "Entfernung_km"]].dropna()
gruppen = df_filtered.groupby(["PersonenID", "Datum"])
anzahl_wege_pro_tag = gruppen.size()
tagesdistanzen = gruppen["Entfernung_km"].sum()
durchschnittliche_wegelänge = gruppen["Entfernung_km"].mean()

ergebnis_df = pd.DataFrame({
    "Anzahl Wege pro Tag": anzahl_wege_pro_tag,
    "Tagesdistanz (km)": tagesdistanzen,
    "Durchschnittliche Wegelänge (km)": durchschnittliche_wegelänge
}).reset_index()

gesamt_ergebnis = pd.DataFrame({
    "Kennzahl": [
        "Durchschnittliche Wegelänge pro Tag (km)",
        "Durchschnittliche Tagesdistanz (km)",
        "Durchschnittliche Anzahl Wege pro Tag"
    ],
    "Wert": [
        round(ergebnis_df["Durchschnittliche Wegelänge (km)"].mean(), 2),
        round(ergebnis_df["Tagesdistanz (km)"].mean(), 2),
        round(ergebnis_df["Anzahl Wege pro Tag"].mean(), 2)
    ]
})

gesamt_ergebnis.to_csv(WEGE_DISTANZEN_CSV_PATH, index=False, encoding='utf-8-sig')
gesamt_ergebnis.to_excel(WEGE_DISTANZEN_XLSX_PATH, index=False)
format_excel(WEGE_DISTANZEN_XLSX_PATH)

# === 8. Wohnviertel Auswertung
df_wohnviertel = df.dropna(subset=["PersonenID", "Wohnviertel"])
wohnviertel_unique = df_wohnviertel.drop_duplicates(subset=["PersonenID"])
wohnviertel_counts = wohnviertel_unique["Wohnviertel"].value_counts().reset_index()
wohnviertel_counts.columns = ["Wohnviertel", "Anzahl Personen"]

gesamt = wohnviertel_counts["Anzahl Personen"].sum()
wohnviertel_counts = pd.concat([
    wohnviertel_counts,
    pd.DataFrame({"Wohnviertel": ["Alle Wohnviertel"], "Anzahl Personen": [gesamt]})
], ignore_index=True)

wohnviertel_counts.to_csv(WOHNVIERTEL_CSV_PATH, index=False, encoding='utf-8-sig')
wohnviertel_counts.to_excel(WOHNVIERTEL_XLSX_PATH, index=False)
format_excel(WOHNVIERTEL_XLSX_PATH)

# === 9. Haushaltstypen Auswertung
df_haushalt = df.dropna(subset=["PersonenID", "Haushaltstyp"])
haushalt_unique = df_haushalt.drop_duplicates(subset=["PersonenID"])
haushalt_counts = haushalt_unique["Haushaltstyp"].value_counts(normalize=True).reset_index()
haushalt_counts.columns = ["Haushaltstyp", "Prozentuale Verteilung"]
haushalt_counts["Prozentuale Verteilung"] = (haushalt_counts["Prozentuale Verteilung"] * 100).round(2)

haushalt_counts.to_csv(HAUSHALTSTYPEN_CSV_PATH, index=False, encoding='utf-8-sig')
haushalt_counts.to_excel(HAUSHALTSTYPEN_XLSX_PATH, index=False)
format_excel(HAUSHALTSTYPEN_XLSX_PATH)

# === 10. Berufe Auswertung
df_beruf = df.dropna(subset=["PersonenID", "Beruf"])
beruf_unique = df_beruf.drop_duplicates(subset=["PersonenID"])
beruf_counts = beruf_unique["Beruf"].value_counts(normalize=True).reset_index()
beruf_counts.columns = ["Beruf", "Prozentuale Verteilung"]
beruf_counts["Prozentuale Verteilung"] = (beruf_counts["Prozentuale Verteilung"] * 100).round(2)

beruf_counts.to_csv(BERUF_CSV_PATH, index=False, encoding='utf-8-sig')
beruf_counts.to_excel(BERUF_XLSX_PATH, index=False)
format_excel(BERUF_XLSX_PATH)

# === 11. Geschlechter Auswertung
df_geschlecht = df.dropna(subset=["PersonenID", "Geschlecht"])
geschlecht_unique = df_geschlecht.drop_duplicates(subset=["PersonenID"])
geschlecht_counts = geschlecht_unique["Geschlecht"].value_counts(normalize=True).reset_index()
geschlecht_counts.columns = ["Geschlecht", "Prozentuale Verteilung"]
geschlecht_counts["Prozentuale Verteilung"] = (geschlecht_counts["Prozentuale Verteilung"] * 100).round(2)

geschlecht_counts.to_csv(GESCHLECHTER_CSV_PATH, index=False, encoding='utf-8-sig')
geschlecht_counts.to_excel(GESCHLECHTER_XLSX_PATH, index=False)
format_excel(GESCHLECHTER_XLSX_PATH)

# === 12. Durchschnittsalter Auswertung
df_alter = df.dropna(subset=["PersonenID", "Alter"])
alter_unique = df_alter.drop_duplicates(subset=["PersonenID"])
durchschnittsalter = alter_unique["Alter"].mean().round(2)

alter_df = pd.DataFrame({"Durchschnittsalter": [durchschnittsalter]})

alter_df.to_csv(ALTER_CSV_PATH, index=False, encoding='utf-8-sig')
alter_df.to_excel(ALTER_XLSX_PATH, index=False)
format_excel(ALTER_XLSX_PATH)

# === 13. Anzahl aufgezeichneter Wege
df_wege = df.dropna(subset=["PersonenID"])
anzahl_wege = df_wege.shape[0]
anzahl_wege_df = pd.DataFrame({"Anzahl aufgezeichneter Wege": [anzahl_wege]})

anzahl_wege_df.to_csv(WEGE_ANZAHL_CSV_PATH, index=False, encoding='utf-8-sig')
anzahl_wege_df.to_excel(WEGE_ANZAHL_XLSX_PATH, index=False)
format_excel(WEGE_ANZAHL_XLSX_PATH)

# === 14. Anzahl befragter Personen
df_personen = df.dropna(subset=["PersonenID"])
anzahl_personen = df_personen["PersonenID"].nunique()
anzahl_personen_df = pd.DataFrame({"Anzahl befragter Personen": [anzahl_personen]})

anzahl_personen_df.to_csv(PERSONEN_ANZAHL_CSV_PATH, index=False, encoding='utf-8-sig')
anzahl_personen_df.to_excel(PERSONEN_ANZAHL_XLSX_PATH, index=False)
format_excel(PERSONEN_ANZAHL_XLSX_PATH)

# === 15. Spitzenstunde Auswertung
df_spitze = df.dropna(subset=["Startzeit"])
df_spitze["Stunde"] = df_spitze["Startzeit"].dt.hour

stunde_counts = df_spitze["Stunde"].value_counts().sort_values(ascending=False)
spitzenstunde = stunde_counts.idxmax()

zeitfenster = f"{spitzenstunde:02d}:00 Uhr bis {spitzenstunde + 1:02d}:00 Uhr"

spitzenstunde_df = pd.DataFrame({"Spitzenstunde": [zeitfenster]})

spitzenstunde_df.to_csv(SPITZENSTUNDE_CSV_PATH, index=False, encoding='utf-8-sig')
spitzenstunde_df.to_excel(SPITZENSTUNDE_XLSX_PATH, index=False)
format_excel(SPITZENSTUNDE_XLSX_PATH)

print("✅ Alle Dateien erfolgreich erstellt und gespeichert.")